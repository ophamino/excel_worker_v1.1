from datetime import datetime

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from logic.const import MAIN_DIR, MONTH_LIST, DEPARTAMENT
from logic.utils import clean_directory


class BicuCalculate:
    
    def serialize_static(self):
        print("[INFO] Сериализуются данные Реестра Бику")
        path = f"{MAIN_DIR}/Реестровая база данных/Реестр БИКУ/Реестр БИКУ.xlsx"
        file = load_workbook(path, data_only=True)
        sheet = file.worksheets[0]
        data = {}
        
        for row in range(5, sheet.max_row + 1):
            
            if sheet.cell(row=row, column=47).value == "Активен": 
                data[sheet.cell(row=row, column=3).value] = {
                    "A": 1,
                    "B": sheet.cell(row=row, column=2).value,
                    "C": sheet.cell(row=row, column=3).value,
                    "D": sheet.cell(row=row, column=4).value,
                    "E": sheet.cell(row=row, column=5).value,
                    "F": str(sheet.cell(row=row, column=8).value),
                    "G": sheet.cell(row=row, column=9).value,
                    "H": sheet.cell(row=row, column=10).value,
                    "I": sheet.cell(row=row, column=13).value,
                    "J": sheet.cell(row=row, column=14).value,
                    "K": sheet.cell(row=row, column=15).value,
                    "L": sheet.cell(row=row, column=20).value,
                    "M": sheet.cell(row=row, column=21).value,
                    "N": sheet.cell(row=row, column=22).value,
                    "O": sheet.cell(row=row, column=23).value,
                    "U": sheet.cell(row=row, column=46).value,
                    "X": sheet.cell(row=row, column=27).value,
                    "Y": sheet.cell(row=row, column=28).value,
                    "Z": sheet.cell(row=row, column=29).value,
                    "AA": sheet.cell(row=row, column=32).value,
                    "AB": sheet.cell(row=row, column=45).value,
                }
        return data
    
    def serialize_svod(self, month: str):
        print("[INFO] Сериализуются данные Сводной ведомости")
        path = f"{MAIN_DIR}/Сводный баланс энергопотребления/Сводный баланс {datetime.now().year}/Сводная ведомость БИКУ/Сводная ведомость БИКУ.xlsx"
        file = load_workbook(path, data_only=True)
        sheet = file[MONTH_LIST[month - 2]]
        data = {}
        
        for row in range(6, sheet.max_row + 1):
            data[sheet.cell(row=row, column=3).value] = {
                "P": sheet.cell(row=row, column=17).value,
                "Q": "",
                "R": "",
                "S": "",
                "T": "",
                "V": "",
                "W": "",
            }
        return data
    
    
    def get_keys(self, static: dict[str, dict], svod: dict[str, dict]):
        keys = []
        for value in svod.values():
            [keys.append(element) for element in value.keys()]
            break
        for value in static.values():
            [keys.append(element) for element in value.keys()]
            break
        keys = sorted(sorted(keys), key=len)
        return keys
    
    def insert_formula(self, sheet: Worksheet):
        for row in range(6, sheet.max_row + 1):
            sheet.cell(row=row, column=1).value = row - 7
            sheet["R{0}".format(row)] = "=IF(Q{0},Q{0}-P{0},0)".format(row)
            sheet["S{0}".format(row)] = "=R{0}*O{0}".format(row)
            sheet["W{0}".format(row)] = "=S{0}+T{0}+U{0}+V{0}".format(row)
            
            
        letter_list = ["R", "S", "T", "U", "V", "W"]
            
        for letter in letter_list:
            sheet[f"{letter}4"] = f'=SUMIFS({letter}8:{letter}{sheet.max_row},$E8:$E{sheet.max_row},"Прием электроэнергии")'
            sheet[f"{letter}5"] = f'=SUMIFS({letter}8:{letter}{sheet.max_row},$E8:$E{sheet.max_row},"Передача электроэнергии")'
            sheet[f"{letter}6"] = f'={letter}4-{letter}5'

    
    def format_data(self, month: int):
        static = self.serialize_static()
        svod = self.serialize_svod(month)
        keys = self.get_keys(svod, static)
        clean_directory(f"{MAIN_DIR}/Шаблоны расчетных ведомостей/РВ БИКУ")
        print("[INFO] Вставляем формулы")
        for departament_id, name in DEPARTAMENT.items():
            file = load_workbook('./template/bicu_v.xlsx')
            sheet = file.worksheets[0]
            for consumer_id, elements in static.items():
                if consumer_id[2:4] == departament_id:
                    row = []
                    for key in keys:
                        if key in elements.keys():
                            row.append(elements[key])
                        if key in svod[consumer_id].keys():
                            row.append(svod[consumer_id][key])
                    sheet.append([value if value else "" for value in row])
            self.insert_formula(sheet)
            file.save(f"{MAIN_DIR}/Шаблоны расчетных ведомостей/РВ БИКУ/РВ БИКУ {name} {datetime.now().year} {MONTH_LIST[month - 1]}.xlsx")
        print("[INFO] ГОТОВО!")