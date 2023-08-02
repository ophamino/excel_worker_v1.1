import os
from datetime import datetime

from openpyxl.worksheet.worksheet import Worksheet

from logic.utils.excel_extends import open_sheet, load_workbook
from logic.const import MAIN_DIR, MONTH_LIST


class BicuComparer:
    
    def collect_files(self, month: int):
        path = f"{MAIN_DIR}/Сводный баланс энергопотребления/Сводный баланс {datetime.now().year}/Сводная ведомость БИКУ/"
        file_path = f"{MAIN_DIR}/Сводный баланс энергопотребления/Сводный баланс {datetime.now().year}/Сводная ведомость БИКУ/Сводная ведомость БИКУ.xlsx"
        
        if not os.path.exists(file_path):
            file = load_workbook('./template/bicu.xlsx')
            file.save(file_path)
        
        file = load_workbook(file_path)
        sheet = open_sheet(file, month)
        
        try:
            file_names = os.listdir(f"{path}{MONTH_LIST[month - 1]}")
            for file_name in file_names:
                rv_file = load_workbook(f"{path}{MONTH_LIST[month - 1]}/{file_name}")
                rv_sheet = rv_file.worksheets[0]
                flag = 8
                for row in rv_sheet.iter_rows(min_row=8, values_only=True):
                    sheet.append(row)
                    sheet.cell(row=flag, column=14).value = str(sheet.cell(row=flag, column=14).value)
                    flag += 1
            self.insert_formula(sheet)
            file.save(file_path)
        except Exception as error:
            print(f"Файл отсутсвует, проверьте директорию {path}{MONTH_LIST[month - 1]}>")
        
    def insert_formula(self, sheet: Worksheet):
        for row in range(8, sheet.max_row + 1):
            sheet.cell(row=row, column=23).value = str(sheet.cell(row=row, column=23).value)
            sheet.cell(row=row, column=1).value = row - 7
            sheet["R{0}".format(row)] = "=Q{0}-P{0}".format(row)
            sheet["S{0}".format(row)] = "=R{0}*O{0}".format(row)
            sheet["W{0}".format(row)] = "=S{0}+T{0}+U{0}+V{0}".format(row)
            
            
        letter_list = ["R", "S", "T", "U", "V", "W"]
            
        for letter in letter_list:
            sheet[f"{letter}4"] = f'=SUMIFS({letter}8:{letter}{sheet.max_row},$E8:$E{sheet.max_row},"Прием электроэнергии")'
            sheet[f"{letter}5"] = f'=SUMIFS({letter}8:{letter}{sheet.max_row},$E8:$E{sheet.max_row},"Передача электроэнергии")'
            sheet[f"{letter}6"] = f'={letter}4-{letter}5'
