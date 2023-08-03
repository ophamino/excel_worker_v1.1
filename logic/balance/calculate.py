from typing import Any
import os
from datetime import datetime

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet

from logic.const import MAIN_DIR, MONTH_LIST
from logic.utils.excel_extends import open_sheet, open_excel


class Balance:
    """
    Класс для формирования сводного баланса
    """

    def serialize_network(self) -> dict[str, dict[str, str]]:
        data = {}
        
        path = f"{MAIN_DIR}/Реестровая база данных/Структура электросети/Свод ОЭСХ.xlsx"
        file = load_workbook(path, data_only=True)
        sheet = file.worksheets[0]
        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row=row, column=1).value:
                data[sheet.cell(row=row, column=1).value] = {
                    "name": sheet.cell(row=row, column=2).value,
                    "consumption": 0,
                    "reception": 0,
                    "transmission": 0,
                    "balance": 0,
                    "waste": 0,
                    "foreign_key": sheet.cell(row=row, column=3).value,
                    "foreign_key_name": sheet.cell(row=row, column=4).value,
                }
        return data
        
    
    def serialize_bicu(self, month: str | int) -> dict[str, dict[str, str]]:
        data = {}
        year, month = datetime.now().year, MONTH_LIST[month - 1]
        files_path = f"{MAIN_DIR}/Сводный баланс энергопотребления/Сводный баланс {year}/Сводная ведомость БИКУ/РВ БИКУ {month}/"
        for file in os.listdir(files_path):
            path = files_path + file
            book = load_workbook(path, data_only=True)
            sheet = book.worksheets[0]
            for row in range(8, sheet.max_row + 1):
                data[sheet.cell(row=row, column=3).value] = {
                    "status": sheet.cell(row=row, column=5).value,
                    "expenses": sheet.cell(row=row, column=23).value
                }
        return data
    
    
    def serialize_consumers(self, month: str | int, status: str) -> dict[str, dict[str, str]]:
        data = {}
        year, month = datetime.now().year, MONTH_LIST[month - 1]
        files_path = f"{MAIN_DIR}/Сводный баланс энергопотребления/Сводный баланс {year}/Сводная ведомость потребителей/РВ Потребителей {month}/РВ {status} потребителей/"
        for file in os.listdir(files_path):
            path = files_path + file
            book = load_workbook(path, data_only=True)
            sheet = book.worksheets[0]
            for row in range(6, sheet.max_row + 1):
                data[sheet.cell(row=row, column=3).value] = {
                    "name": sheet.cell(row=row, column=8).value,
                    "expenses": sheet.cell(row=row, column=29).value,
                    "foreign_key": sheet.cell(row=row, column=39).value
                }
        return data
        
        

    def serialize_balance(self, month: str | int) -> dict[str, dict[str, str]]:
        data = self.serialize_network()
        bicu_data = self.serialize_bicu(month)
        consumers_data = self.serialize_consumers(month, "Бытовых") | self.serialize_consumers(month, "Коммерческих")
        
        for key in bicu_data.keys():
            if bicu_data[key]["status"] == "Прием электроэнергии":
                data[key]["reception"] += bicu_data[key]["expenses"]
            if bicu_data[key]["status"] == "Передача электроэнергии":
                data[key]["transmission"] += bicu_data[key]["expenses"]
        
        for key in consumers_data.keys():
            foreign_key = consumers_data[key]["foreign_key"]
            data[foreign_key]["consumption"] += consumers_data[key]["expenses"]
        
        keys = sorted(data.keys())[::-1]
        
        for key in keys:
            foreign_key = data[key]["foreign_key"]
            if foreign_key:
                data[foreign_key]["reception"] += data[key]["reception"]
                data[foreign_key]["transmission"] += data[key]["transmission"]
                data[foreign_key]["consumption"] += data[key]["consumption"]
                data[foreign_key]["balance"] = data[foreign_key]["reception"] - data[foreign_key]["transmission"]
                data[foreign_key]["waste"] = data[foreign_key]["balance"] - data[foreign_key]["consumption"]
        
        return data
        
    def create_balance(self, month: str | int) -> None:
        data = self.serialize_balance(month)
        path = f"{MAIN_DIR}/Сводный баланс энергопотребления/Сводный баланс {datetime.now().year}/Сводный баланс 2023.xlsx"
        file = open_excel(path)
        sheet = open_sheet(file, month=month)
        
        sheet.append(["№", "Идентификатор", "Наименование", "Вход", "Выход", "Сальдо переток", "Полезный отпуск", "Потери", "Процент потерь"])
        number = 1
        for key in data.keys():
            try:
                percent = round(data[key]["waste"] / data[key]["balance"] * 100, 2)
            except ZeroDivisionError:
                percent = 0
            if data[key]["balance"] == 0:
                data[key]["balance"] = ""
            sheet.append(
                [
                    number, key, data[key]["name"], data[key]["reception"],
                    data[key]["transmission"], data[key]["balance"], data[key]["consumption"],
                    data[key]["waste"], percent
                ]
            )
            number += 1
        file.save(path)
