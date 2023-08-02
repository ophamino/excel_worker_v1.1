import os
from datetime import datetime

from openpyxl.worksheet.worksheet import Worksheet

from logic.utils.excel_extends import open_sheet, load_workbook
from logic.const import MAIN_DIR, MONTH_LIST


class ConsumersComparer:
    
    def collect_files(self, month: int, file_status: str):
        path = f"{MAIN_DIR}/Сводный баланс энергопотребления/Сводный баланс {datetime.now().year}/Сводная ведомость потребителей/"
        file_path = f"{MAIN_DIR}/Сводный баланс энергопотребления/Сводный баланс {datetime.now().year}/Сводная ведомость потребителей/Сводная ведомость {file_status} потребителей.xlsx"
        dir_path = f"{MAIN_DIR}/Сводный баланс энергопотребления/Сводный баланс {datetime.now().year}/Сводная ведомость потребителей/{MONTH_LIST[month - 1]}/РВ {file_status} потребителей"
        
        if not os.path.exists(file_path):
            file = load_workbook('./template/svod.xlsx')
            file.save(file_path)
        
        file = load_workbook(file_path)
        sheet = open_sheet(file, month)
        
        try:
            file_names = os.listdir(dir_path)
            for file_name in file_names:
                rv_file = load_workbook(f"{dir_path}/{file_name}")
                rv_sheet = rv_file.worksheets[0]
                flag = 6
                flag_2 = 6
                for row in rv_sheet.iter_rows(min_row=6, values_only=True):
                    if row[0]:
                        sheet.append(row)
                        sheet.cell(row=flag, column=6).value = str(sheet.cell(row=flag, column=6).value)
                        sheet.cell(row=flag, column=19).value = str(sheet.cell(row=flag, column=19).value)
                        flag += 1
            self.insert_formula(sheet)
            file.save(file_path)
        except Exception as error:
            print(f"Файл отсутсвует, проверьте директорию {path}{MONTH_LIST[month - 1]}>")
    
    def collect_total_files(self, month: int):
        path = f"{MAIN_DIR}/Сводный баланс энергопотребления/Сводный баланс {datetime.now().year}/Сводная ведомость потребителей/"
        
        if not os.path.exists(f"{path}\Сводная ведомость потребителей.xlsx"):
            file = load_workbook('template/svod.xlsx')
            file.save(f"{path}\Сводная ведомость потребителей.xlsx")
        file = load_workbook(f"{path}\Сводная ведомость потребителей.xlsx")
        sheet = file[MONTH_LIST[month - 1]]
        try:
            comerce = load_workbook(f"{path}\Сводная ведомость Коммерческих потребителей.xlsx", data_only=True)[MONTH_LIST[month - 1]]
            individual = load_workbook(f"{path}\Сводная ведомость Бытовых потребителей.xlsx", data_only=True)[MONTH_LIST[month - 1]]
        except Exception as e:
            print(e)
        
        for row in individual.iter_rows(min_row=6, values_only=True):
            sheet.append(row)
        for row in comerce.iter_rows(min_row=6, values_only=True):
            sheet.append(row)
        self.insert_formula(sheet)
        file.save(f"{path}\Сводная ведомость потребителей.xlsx")
        
    def insert_formula(self, sheet: Worksheet):
        for row in range(6, sheet.max_row + 1):
            sheet.cell(row=row, column=29).value = str(sheet.cell(row=row, column=29).value)
            sheet.cell(row=row, column=1).value = row - 5
            sheet.cell(row=row, column=23).value = "=V{0}-U{0}".format(row)
            sheet.cell(row=row, column=24).value = "=W{0}*T{0}".format(row)
            sheet.cell(row=row, column=25).value = "=IF(ISBLANK($AL${0}),0,ROUND(($X${0}*$AL${0}),0))".format(row)
            sheet.cell(row=row, column=29).value = "=X{0}+Y{0}+Z{0}+AA{0}+AB{0}".format(row)
        sheet.cell(row=4, column=23).value = "=SUM(W6:W{0})".format(sheet.max_row)
        sheet.cell(row=4, column=24).value = "=SUM(X6:X{0})".format(sheet.max_row)
        sheet.cell(row=4, column=25).value = "=SUM(Y6:Y{0})".format(sheet.max_row)
        sheet.cell(row=4, column=26).value = "=SUM(Z6:Z{0})".format(sheet.max_row)
        sheet.cell(row=4, column=27).value = "=SUM(AA6:AA{0})".format(sheet.max_row)
        sheet.cell(row=4, column=28).value = "=SUM(AB6:AB{0})".format(sheet.max_row)
        sheet.cell(row=4, column=29).value = "=SUM(AC6:AC{0})".format(sheet.max_row)