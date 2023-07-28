from datetime import datetime

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from logic.const import MAIN_DIR, MONTH_LIST, DEPARTAMENT


class Calculate:
    
    def serialize_static(self, status: str):
        print("[INFO] Сериализуются данные Реестра потребителей")
        path = f"{MAIN_DIR}/Реестровая база данных/Реестр потребителей/Реестр потребителей.xlsx"
        file = load_workbook(path, data_only=True)
        sheet = file.worksheets[0]
        data = {}
        if status == "Коммерческих":
            status = "cc"
        if status == "Бытовых":
            status = "ch"
        
        for row in range(9, sheet.max_row + 1):
            if sheet.cell(row=row, column=2).value == status and sheet.cell(row=row, column=55).value == "Активен": 
                data[sheet.cell(row=row, column=4).value] = {
                    "A": 1,
                    "B": sheet.cell(row=row, column=3).value,
                    "C": sheet.cell(row=row, column=4).value,
                    "D": sheet.cell(row=row, column=5).value,
                    "E": sheet.cell(row=row, column=6).value,
                    "F": str(sheet.cell(row=row, column=7).value),
                    "G": sheet.cell(row=row, column=8).value,
                    "H": sheet.cell(row=row, column=9).value,
                    "I": sheet.cell(row=row, column=12).value,
                    "J": sheet.cell(row=row, column=13).value,
                    "K": sheet.cell(row=row, column=14).value,
                    "L": sheet.cell(row=row, column=20).value,
                    "M": sheet.cell(row=row, column=21).value,
                    "N": sheet.cell(row=row, column=22).value,
                    "O": sheet.cell(row=row, column=23).value,
                    "P": sheet.cell(row=row, column=24).value,
                    "Q": sheet.cell(row=row, column=27).value,
                    "R": sheet.cell(row=row, column=28).value,
                    "S": str(sheet.cell(row=row, column=29).value),
                    "T": sheet.cell(row=row, column=30).value,
                    "AM": sheet.cell(row=row, column=54).value,
                }
        return data
    
    def serialize_svod(self, month: str, status: str):
        print("[INFO] Сериализуются данные Сводной ведомости")
        path = f"{MAIN_DIR}/Сводный баланс энергопотребления/Сводный баланс {datetime.now().year}/Сводная ведомость потребителей/Сводная ведомость {status} потребителей.xlsx"
        file = load_workbook(path, data_only=True)
        sheet = file[MONTH_LIST[month - 1]]
        data = {}
        
        for row in range(6, sheet.max_row + 1):
            data[sheet.cell(row=row, column=3).value] = {
                "U": sheet.cell(row=row, column=22).value,
                "V": sheet.cell(row=row, column=22).value,
                "W": "",
                "X": "",
                "Y": "",
                "Z": sheet.cell(row=row, column=26).value,
                "AA": "",
                "AB": "",
                "AC": "",
                "AD": sheet.cell(row=row, column=30).value,
                "AE": "",
                "AF": "",
                "AG": "",
                "AH": "",
                "AI": sheet.cell(row=row, column=35).value,
                "AJ": sheet.cell(row=row, column=36).value,
                "AK": sheet.cell(row=row, column=37).value,
                "AL": sheet.cell(row=row, column=38).value,
                "AN": "",
                "AO": "",
                "AP": sheet.cell(row=row, column=40).value,
                "AQ": sheet.cell(row=row, column=41).value,
            }
        return data
    
    
    def get_keys(self, static: dict[str, dict], svod: dict[str, dict]):
        keys = []
        for value in svod.values():
            [keys.append(element) for element in value.keys()]
            break
        for value in static.values():
            [keys.append(element) for element in value.keys()]
            break
        keys = sorted(sorted(keys), key=len)
        return keys
    
    def insert_formula(self, sheet: Worksheet):
        for row in range(6, sheet.max_row+1):
            sheet.cell(row=row, column=23).value = "=V{0}-U{0}".format(row)
            sheet.cell(row=row, column=24).value = "=W{0}*T{0}".format(row)
            sheet.cell(row=row, column=25).value = "=IF(ISBLANK($AL${0}),0,ROUND(($X${0}*$AL${0}),0))".format(row)
            sheet.cell(row=row, column=29).value = "=X{0}+Y{0}+Z{0}+AA{0}+AB{0}".format(row)
        sheet.cell(row=4, column=23).value = "=SUM(W6:W{0})".format(sheet.max_row + 1)
        sheet.cell(row=4, column=24).value = "=SUM(X6:X{0})".format(sheet.max_row + 1)
        sheet.cell(row=4, column=25).value = "=SUM(Y6:Y{0})".format(sheet.max_row + 1)
        sheet.cell(row=4, column=26).value = "=SUM(Z6:Z{0})".format(sheet.max_row + 1)
        sheet.cell(row=4, column=27).value = "=SUM(AA6:AA{0})".format(sheet.max_row + 1)
        sheet.cell(row=4, column=28).value = "=SUM(AB6:AB{0})".format(sheet.max_row + 1)
        sheet.cell(row=4, column=29).value = "=SUM(AC6:AC{0})".format(sheet.max_row + 1)
    
    
    def format_data(self, month: int, status: str):
        static = self.serialize_static(status)
        svod = self.serialize_svod(month, status)
        keys = self.get_keys(svod, static)
        print("[INFO] Вставляем формулы")
        for departament_id, name in DEPARTAMENT.items():
            file = load_workbook('./template/rv.xlsx')
            sheet = file.worksheets[0]
            for consumer_id, elements in static.items():
                if consumer_id[2:4] == departament_id:
                    row = []
                    for key in keys:
                        if key in elements.keys():
                            row.append(elements[key])
                        if key in svod[consumer_id].keys():
                            row.append(svod[consumer_id][key])
                    sheet.append(row)
            self.insert_formula(sheet)
            file.save(f"{MAIN_DIR}/Шаблоны расчетных ведомостей/РВ {status} потребителей/РВ {name} {datetime.now().year} {MONTH_LIST[month - 1]}.xlsx")
        print("[INFO] ГОТОВО!")