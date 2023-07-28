import os
from datetime import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from logic.const import MONTH_LIST


def open_excel(path: str, data_only: bool = False, read_only: bool = False) -> Workbook:
        """
        Открывает файл, если его не существует - создаёт и открывает\n
        Args: path (str): путь к файлу, где должен находиться файл
        """
        if not os.path.exists(path):
            workbook = Workbook()
            workbook.save(path)
        return load_workbook(path, data_only=data_only, read_only=read_only)
    
def create_months_sheet_if_not_exists(workbook: Workbook, month: str | int | None = None) -> None:
    """
    Создает месячный лист, если его не существует

    Args:
        workbook (Workbook): Документ excel, который будет проверяться\n\n
        month (str | int | None): Месячный лист, который должен находится в файле:\n
        Если значение `str`, то лист создаётся по названию;\n
        Если знаение `int`, то лист создается по номеру согласно порядку;\n
        Если значение `None`, то создается лист за текущий месяц.\n
        По умолчанию `None`.
    """

    sheetnames = workbook.sheetnames

    def create(workbook: Workbook, month: str | int, sheetnames: list[str]):
        if month not in sheetnames:
            workbook.create_sheet(month)

    if isinstance(month, str):
        if month in MONTH_LIST:
            create(workbook, month, sheetnames)
        else:
            raise ValueError(f"Указано неправильное название месяца: {month}")

    if isinstance(month, int):
        if 1 <= month <= len(MONTH_LIST):
            month = MONTH_LIST[month - 1]
            create(workbook, month, sheetnames)
        else:
            raise ValueError(f"Указан неправильный номер месяца: {month}")

    if month is None:
        month = datetime.now().month
        month = MONTH_LIST[month - 1]
        create(workbook, month, sheetnames)
        
            
def open_sheet( workbook: Workbook, month: str | int) -> Worksheet:
    """
    Открывает лист по номеру или названию месяца

    Args:
        workbook (Workbook): Excel файл, лист которого нужно открыть\n
        month (str | int): Номер или название месяца

    Returns:
        Worksheet: Лист Excel файла
    """        
    create_months_sheet_if_not_exists(workbook, month)
    if isinstance(month, int):
        month = MONTH_LIST[month - 1]
    return workbook[month]
