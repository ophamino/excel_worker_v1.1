from matplotlib import pyplot as plt
import numpy as np

from logic.balance.calculate import Balance
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import PatternFill, Alignment, Font

class BalanceAnalytics:
    
    model = Balance()
    
    titles_font = Font(
        size=28,
        bold=True,
    )
    
    alignment = Alignment(
        horizontal="center",
        vertical="center"
    )
    
    def get_data(self, month: str | int):
        data = self.model.serialize_balance(month)
        return data
    
    def write_balance_pie(self, data: dict):
        labels = list([value["name"] for key, value in data.items() if key != "DS"])
        slices = list([abs(value["balance"]) for key, value in data.items() if key != "DS" if value["balance"] != 0])
        if not slices:
            slices += [1] * 5
        if not data["DS"]["balance"]:
            data["DS"]["balance"] = 5
        percent = list([f'{round(i / abs(data["DS"]["balance"]) * 100, 1)}%' for i in slices if i != 0])
        fig, ax = plt.subplots()
        ax.pie(slices,labels=percent , wedgeprops={"edgecolor": "black", "linewidth": 1})
        ax.legend(labels, loc=2)
        plt.pie([1], colors="w", radius=0.6)
        plt.title("Сальдо переток")
        plt.savefig("./images/balance_pie.png")
        
    def write_waste_pie(self, data):
        slices = list([abs(value["waste"]) for key, value in data.items() if key != "DS" if value["waste"] != 0])
        if not slices:
            slices += [1] * 5
        if not data["DS"]["waste"]:
            data["DS"]["waste"] = 5
        percent = list([f'{round(i / abs(data["DS"]["waste"]) * 100, 1)}%' for i in slices if i != 0])
        fig, ax = plt.subplots()
        ax.pie(slices, labels=percent, wedgeprops={"edgecolor": "black", "linewidth": 1})
        plt.pie([1], colors="w", radius=0.6)
        plt.title("Потери")
        plt.savefig("./images/waste_pie.png")
    
    def write_consumption_pie(self, data):
        slices = list([abs(value["consumption"]) for key, value in data.items() if key != "DS" if value["consumption"] != 0])
        if not slices:
            slices += [1] * 5
        if not data["DS"]["consumption"]:
            data["DS"]["consumption"] = 5
        percent = list([f'{round(i / abs(data["DS"]["consumption"]) * 100, 1)}%' for i in slices if i != 0])
        fig, ax = plt.subplots()
        ax.pie(slices, labels=percent, wedgeprops={"edgecolor": "black", "linewidth": 1})
        plt.pie([1], colors="w", radius=0.6)
        plt.title("Полезный отпуск")
        plt.savefig("./images/consumption_pie.png")
    
    def create_all_pie(self, sheet: Worksheet, data):
        self.write_waste_pie(data)
        self.write_balance_pie(data)
        self.write_consumption_pie(data)
        
        balance_pie = Image("./images/balance_pie.png")
        balance_pie.anchor = 'A9'
        sheet.add_image(balance_pie)
        
        consumption = Image("./images/consumption_pie.png")
        consumption.anchor = 'K9'
        sheet.add_image(consumption)
        
        waste_pie = Image("./images/waste_pie.png")
        waste_pie.anchor = 'U9'
        sheet.add_image(waste_pie)
    
    def write_bars(self, data: dict):
        fig, ax = plt.subplots()
        slices = sorted(list([abs(value["waste"]) for key, value in data.items() if key != "DS" if value["waste"] != 0]))
        labels = list([value["name"] for value in reversed(sorted(data.values(), key=lambda x: x["waste"])) if value["foreign_key"] is not None])

        ax.barh(labels, slices, xerr=min(slices), align="center")
        plt.title('Потери по структурным подразделениям')
        plt.ylabel('Структурные подразделения')
        plt.xlabel('Потери (в млн.)')
        plt.savefig("./images/hbar.png")

    def write_departamnents_bars(self, data):
        names = []
        slices = {
            "Сальдо переток": [],
            "Полезный отпуск": [],
            "Потери": [],
        }
        
        for key, value in data.items():
            if key != "DS":
                names.append(value["name"])
                max_num = max([int(value["balance"] / 1000), int(value["consumption"] / 1000), int(abs(value["waste"] / 1000))])
                if value["balance"]:
                    slices["Сальдо переток"].append(max_num / int(value["consumption"] / 1000) * 100)
                else:
                    slices["Сальдо переток"].append(0)
                if value["consumption"]:
                    slices["Полезный отпуск"].append(max_num / int(value["consumption"] / 1000) * 100)
                else:
                    slices["Полезный отпуск"].append(0)
                if  value["waste"]:
                    slices["Потери"].append(max_num /  int(abs(value["waste"] / 1000)) * 100)
                else:
                    slices["Потери"].append(0)
                
                
        x = np.arange(len(names))
        width = 0.25
        multiplier = 0

        fig, ax = plt.subplots(layout='constrained', figsize=(3.5*7.3, 3.5 * 1.5))

        for attribute, measurement in slices.items():
            offset = width * multiplier
            rects = ax.bar(x + offset, measurement, width, label=attribute)
            ax.bar_label(rects, padding=3)
            multiplier += 1

        # Add some text for labels, title and custom x-axis tick labels, etc.
        ax.set_ylabel('Показатели (в тыс.)')
        ax.set_title('Индивидуальные показатели по структурным подразделениям')
        ax.set_xticks(x + width, names)
        ax.legend(loc='upper left', ncols=3)
        ax.set_ylim(0, 100)
        plt.rcParams["figure.figsize"] = 10, 20
        # plt.savefig("./images/all_bars.png")

    def create_analytics(self, month):
        data = self.model.serialize_balance(month)
        book = load_workbook(r"C:\Users\user\Desktop\Тест.xlsx")
        sheet = book.worksheets[0]
        
        sheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=40)
        sheet.cell(row=1, column=1).value = 'ООО "ДагЭнерЖи"'
        sheet.cell(row=1, column=1).font = self.titles_font
        sheet.cell(row=1, column=1).alignment = self.alignment
        sheet.cell(row=1, column=1).fill = PatternFill(fgColor="ECEDE8", fill_type="solid")
        change_start_col = 1
        change_end_col = 8
        
        for _ in range(5):
            sheet.merge_cells(start_row=3, start_column=change_start_col, end_row=5, end_column=change_end_col)
            sheet.merge_cells(start_row=6, start_column=change_start_col, end_row=8, end_column=change_end_col)
            change_start_col += 8
            change_end_col += 8
        colors = ["99CCFF", "CCCCCC", "FFFFCC", "CCFFCC", "FF9999"]
        titles = ["Вход", "Выход", "Сальдо переток", "Полезный отпуск", "Потери"]
        for step, title_num in zip(range(1, 40, 8), range(5)):
            fill = PatternFill(fgColor=colors[title_num], fill_type="solid")
            sheet.cell(row=3, column=step).value = titles[title_num]
            sheet.cell(row=3, column=step).font = self.titles_font
            sheet.cell(row=3, column=step).alignment = self.alignment
            sheet.cell(row=3, column=step).fill = fill
        
        attributes = [data["DS"]["reception"], data["DS"]["transmission"],  data["DS"]["balance"], data["DS"]["consumption"], data["DS"]["waste"]]
        for step, attributes_num in zip(range(1, 40, 8), range(5)):
            sheet.cell(row=6, column=step).value = attributes[attributes_num]
            sheet.cell(row=6, column=step).font = self.titles_font
            sheet.cell(row=6, column=step).alignment = self.alignment
        
        self.create_all_pie(sheet, data)
        
        self.write_bars(data=data)
        self.write_departamnents_bars(data=data)

        hbar = Image("./images/hbar.png")
        hbar.anchor = 'AE9'
        sheet.add_image(hbar)
    
        
        book.save("../test/Аналитика баланса электроэнергии/Аналитика Сводного баланса.xlsx")

